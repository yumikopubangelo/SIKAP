import json
import os
import uuid
from datetime import date, datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from ..extensions import db
from ..models import Absensi, Kelas, Laporan, SesiSholat, Siswa, WaktuSholat

REPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "static", "generated")


def _ensure_report_dir() -> None:
    os.makedirs(REPORT_DIR, exist_ok=True)


def _fetch_rekap_data(
    jenis: str,
    filter_id: int | None = None,
    tanggal_mulai: date | None = None,
    tanggal_selesai: date | None = None,
) -> list[dict]:
    query = (
        Absensi.query.options(
            joinedload(Absensi.siswa).joinedload(Siswa.kelas),
            joinedload(Absensi.sesi).joinedload(SesiSholat.waktu_sholat),
        )
        .join(Absensi.siswa)
        .join(Absensi.sesi)
    )

    if jenis == "kelas" and filter_id:
        query = query.filter(Siswa.id_kelas == filter_id)
    elif jenis == "siswa" and filter_id:
        query = query.filter(Absensi.id_siswa == filter_id)
    elif jenis == "sekolah":
        pass

    if tanggal_mulai:
        query = query.filter(SesiSholat.tanggal >= tanggal_mulai)
    if tanggal_selesai:
        query = query.filter(SesiSholat.tanggal <= tanggal_selesai)

    absensi_list = query.order_by(Absensi.timestamp.desc()).all()

    rows = []
    for item in absensi_list:
        rows.append(
            {
                "nama_siswa": item.siswa.nama if item.siswa else "-",
                "nisn": item.siswa.nisn if item.siswa else "-",
                "kelas": item.siswa.kelas.nama_kelas if item.siswa and item.siswa.kelas else "-",
                "waktu_sholat": item.sesi.waktu_sholat.nama_sholat if item.sesi and item.sesi.waktu_sholat else "-",
                "tanggal": item.sesi.tanggal.isoformat() if item.sesi and item.sesi.tanggal else "-",
                "timestamp": item.timestamp.isoformat() if item.timestamp else "-",
                "status": item.status,
                "keterangan": item.keterangan or "-",
            }
        )
    return rows


def _build_pdf(filename: str, rows: list[dict], judul: str) -> str:
    _ensure_report_dir()
    filepath = os.path.join(REPORT_DIR, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=2 * cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(judul, styles["Title"]))
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(f"Digenerate: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    headers = ["No", "Nama Siswa", "NISN", "Kelas", "Waktu Sholat", "Tanggal", "Status", "Keterangan"]
    table_data = [headers]
    for idx, row in enumerate(rows, start=1):
        table_data.append(
            [
                idx,
                row["nama_siswa"],
                row["nisn"],
                row["kelas"],
                row["waktu_sholat"],
                row["tanggal"],
                row["status"],
                row["keterangan"],
            ]
        )

    col_widths = [1.2 * cm, 3.5 * cm, 2.5 * cm, 2 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F5F5F5")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return filepath


def _build_excel(filename: str, rows: list[dict], judul: str) -> str:
    _ensure_report_dir()
    filepath = os.path.join(REPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Kehadiran"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    title_font = Font(bold=True, size=14)

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = judul
    title_cell.font = title_font

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Digenerate: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, size=10)

    headers = ["No", "Nama Siswa", "NISN", "Kelas", "Waktu Sholat", "Tanggal", "Status", "Keterangan"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(rows, start=5):
        ws.cell(row=idx, column=1, value=idx - 4)
        ws.cell(row=idx, column=2, value=row["nama_siswa"])
        ws.cell(row=idx, column=3, value=row["nisn"])
        ws.cell(row=idx, column=4, value=row["kelas"])
        ws.cell(row=idx, column=5, value=row["waktu_sholat"])
        ws.cell(row=idx, column=6, value=row["tanggal"])
        ws.cell(row=idx, column=7, value=row["status"])
        ws.cell(row=idx, column=8, value=row["keterangan"])

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    wb.save(filepath)
    return filepath


def generate_laporan(
    jenis: str,
    format_laporan: str,
    filter_id: int | None = None,
    tanggal_mulai: date | None = None,
    tanggal_selesai: date | None = None,
    user_id: int | None = None,
) -> Laporan:
    rows = _fetch_rekap_data(jenis, filter_id, tanggal_mulai, tanggal_selesai)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = uuid.uuid4().hex[:8]
    base_name = f"laporan_{jenis}_{timestamp}_{unique_id}"

    if format_laporan == "pdf":
        filename = f"{base_name}.pdf"
        judul = f"Laporan Kehadiran - {jenis.title()}"
        filepath = _build_pdf(filename, rows, judul)
    elif format_laporan == "excel":
        filename = f"{base_name}.xlsx"
        judul = f"Laporan Kehadiran - {jenis.title()}"
        filepath = _build_excel(filename, rows, judul)
    else:
        raise ValueError(f"Format tidak didukung: {format_laporan}")

    filter_data = json.dumps(
        {
            "jenis": jenis,
            "filter_id": filter_id,
            "tanggal_mulai": tanggal_mulai.isoformat() if tanggal_mulai else None,
            "tanggal_selesai": tanggal_selesai.isoformat() if tanggal_selesai else None,
        }
    )

    laporan = Laporan(
        jenis=jenis,
        format=format_laporan,
        filter_data=filter_data,
        file_path=filepath,
        generated_by=user_id,
    )
    db.session.add(laporan)
    db.session.commit()
    return laporan
